from pdfminer3.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer3.converter import PDFPageAggregator
from pdfminer3.pdfpage import PDFPage
from pdfminer3.layout import LAParams, LTTextContainer
import openpyxl

import pandas as pd
import math
import os
import csv
import datetime
import io


def pdfminer_config(line_overlap, word_margin, char_margin,line_margin, detect_vertical):
    laparams = LAParams(line_overlap=line_overlap,
                        word_margin=word_margin,
                        char_margin=char_margin,
                        line_margin=line_margin,
                        detect_vertical=detect_vertical)
    resource_manager = PDFResourceManager()
    device = PDFPageAggregator(resource_manager, laparams=laparams)
    interpreter = PDFPageInterpreter(resource_manager, device)
    return (interpreter, device)

def remove_empty_rows_cols(ws):
    # 空白行を削除
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 逆順に走査して空白行を削除（行数が変わるため）
    for row in range(max_row, 0, -1):
        if all([ws.cell(row=row, column=col).value is None for col in range(1, max_col + 1)]):
            ws.delete_rows(row, 1)

    # 空白列を削除
    for col in range(max_col, 0, -1):
        if all([ws.cell(row=row, column=col).value is None for row in range(1, max_row + 1)]):
            ws.delete_cols(col, 1)

# @param bytesio, pdf
# @return bytesio, cvs
def pdf2csv(pdf):
    # print(f"file for {pdf}")
    # work_file = os.path.splitext(pdf_file_name)[0] + '_work.xlsx'
    # excel_file_name = os.path.splitext(pdf_file_name)[0] + '.xlsx'

    list1 = ['','','','','','','','']
    df_x = pd.DataFrame([list1])
    df_x.columns = ['page', 'word', 'x1','x2','y1','y2','width','hight']
    int_page = 0
    ii_index = 0

    with pdf as fp:
        interpreter, device = pdfminer_config(line_overlap=0.1, word_margin=0.1, 
                char_margin=0.1, line_margin=0.3, detect_vertical=True)
        for page in PDFPage.get_pages(fp):
            int_page = int_page + 1
            interpreter.process_page(page)
            layout = device.get_result()
            for lt in layout:
                # LTTextContainerの場合だけ標準出力
                if isinstance(lt, LTTextContainer):
                    df_x.loc[ii_index] = [int_page,'{}'.format(lt.get_text().strip()), lt.x0 , lt.x1 ,\
                    841 - lt.y0 + (int_page - 1) * 841,841 - lt.y1  + (int_page - 1) * 841,lt.width ,lt.height ]
                    ii_index = ii_index + 1

    device.close()

    # x1でソート
    df_s_x = df_x.sort_values(['x1','y2'], ascending=[True,True])
    df_s_x = df_s_x.reset_index(drop=True)
    df_s_x = df_s_x.drop(index = [0, 1, 2])
    # df_s_x.drop(df_s_x.index[[0, 1, 2]])

    # 縦のピッチを計算
    h_min = 100
    for i in range(len(df_s_x)):
        if i > 0:
            if df_s_x.iloc[i-1,2] == df_s_x.iloc[i,2]:
                h_sa = df_s_x.iloc[i,5] - df_s_x.iloc[i-1,5] 
                if h_sa > 1.0 and h_min > h_sa:
                    h_min = h_sa

    # workファイルを書き出し
    # with pd.ExcelWriter(work_file) as writer:
    #     df_s_x.to_excel(writer, sheet_name='sheet1', index=False)

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    j = 1
    width_x = 0
    for i in range(len(df_s_x)):
        y = df_s_x.iloc[i,5] // (math.ceil(h_min*10)/10) + 1
        c1 = ws.cell(row=int(y), column=j)
        if c1.value == None:
            c1.value = df_s_x.iloc[i,1]
        else:
            #列幅調整
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = (df_s_x.iloc[i,2]/5.98- width_x )
            width_x = df_s_x.iloc[i,2]/5.98
            j = j + 1
            c1 = ws.cell(row=int(y), column=j)
            c1.value = df_s_x.iloc[i,1]

    # periodはpandasから
    date = df_s_x[df_s_x["word"].str.startswith("期　　間")].iloc[0, 1]
    date = date.split("～")
    start = date[0][5:]
    end = date[1]
    strformat = "%Y/%m/%d"
    period = (datetime.datetime.strptime(end, strformat) - datetime.datetime.strptime(start, strformat)).days+1
    dateobj = (start, end)

    # 削除はwsから
    ws.delete_rows(1, amount=16)
    remove_empty_rows_cols(ws)
    # wb.save(excel_file_name)

    # ファイル保存
    # with open(os.path.splitext(pdf_file_name)[0]+".csv", "w", newline="", encoding="shift-jis", errors="ignore") as f:
    #     writer = csv.writer(f)
        
    #     for row in ws.iter_rows(values_only=True):
    #         writer.writerow(row)

    # binary
    output = io.StringIO()
    csvwriter = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    for row in ws.iter_rows(values_only=True):
        csvwriter.writerow(row)

    output_s = output.getvalue()
    output.close()

    return output_s, dateobj